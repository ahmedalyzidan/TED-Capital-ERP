import json
import os
import sys
import traceback
from pathlib import Path
from difflib import SequenceMatcher


DOCX_PATH = Path(r"D:/z/Digital onboarding/AAIB_Business_Digital_Onboarding_RFP_VF 18-05-2026_Backup.docx")
XLSX_PATH = Path(r"D:/z/Digital onboarding/AAIB_Digital_Onboarding_Requirements_Comparison.xlsx")
OUT_JSON = Path(r"D:/z/Digital onboarding/header_discovery_diagnostics_relaxed.json")

TARGET = "Aligned BRD Description"
THRESHOLD = 0.2


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def extract_docx_headings(path: Path):
    try:
        from docx import Document
    except Exception as e:
        print(f"python-docx not available: {e}")
        return []

    if not path.exists():
        print(f"DOCX file not found: {path}")
        return []

    headings = []
    try:
        doc = Document(path)
        for p in doc.paragraphs:
            style_name = getattr(p.style, 'name', '')
            if style_name in ("Heading 1", "Heading 2", "Heading 3", "Heading 4"):
                headings.append({"text": p.text.strip(), "style": style_name})
    except Exception as e:
        print(f"Error reading DOCX {path}: {e}")
    return headings


def analyze_excel(path: Path):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"openpyxl not available: {e}")
        return {}

    if not path.exists():
        print(f"XLSX file not found: {path}")
        return {}

    diagnostics = {}
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            # read first 6 rows for context
            first_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                first_rows.append([safe_str(c) for c in row])
                if i >= 6:
                    break

            # print sheet and context
            print(f"\nSheet: {sheetname}")
            max_cols = max((len(r) for r in first_rows), default=0)
            for idx, r in enumerate(first_rows, start=1):
                print(f" Row {idx:>2}:", r)

            # detect header rows: first non-empty row
            header_row_idx = None
            for idx, r in enumerate(first_rows, start=1):
                non_empty = sum(1 for c in r if c and c.strip())
                if non_empty > 0:
                    header_row_idx = idx
                    break

            headers = []
            header_preview = []
            # try multi-row header detection using first two rows
            if header_row_idx is None:
                print("  No header row detected in first 6 rows.")
                diagnostics[sheetname] = []
                continue

            # get header row values
            r1 = first_rows[header_row_idx - 1] if header_row_idx - 1 < len(first_rows) else []
            r2 = first_rows[header_row_idx] if header_row_idx < len(first_rows) else []

            def non_empty_count(row):
                return sum(1 for c in row if c and c.strip())

            # decide if second row is also header-ish
            use_multi = False
            if non_empty_count(r1) >= 1 and non_empty_count(r2) >= max(2, int(0.3 * max(len(r1), len(r2), 1))):
                use_multi = True

            cols = max(len(r1), len(r2))
            for ci in range(cols):
                part1 = safe_str(r1[ci]) if ci < len(r1) else ""
                part2 = safe_str(r2[ci]) if ci < len(r2) else ""
                if use_multi and part2:
                    colname = f"{part1} | {part2}" if part1 else part2
                else:
                    colname = part1
                colname = colname.strip()
                headers.append(colname)
                header_preview.append([part1, part2])

            # compute similarity scores
            sheet_entries = []
            target_norm = TARGET.strip().lower()
            for col, preview in zip(headers, header_preview):
                norm = col.strip().lower()
                score = 0.0
                if norm:
                    score = SequenceMatcher(None, norm, target_norm).ratio()
                entry = {
                    "column": col,
                    "score": round(float(score), 6),
                    "header_row_preview": preview,
                }
                sheet_entries.append(entry)
                mark = "" if score < THRESHOLD else " <-- candidate"
                print(f"  Column: '{col}'  Score: {score:.4f}{mark}")

            diagnostics[sheetname] = sheet_entries

    except Exception as e:
        print(f"Error reading XLSX {path}: {e}")
        traceback.print_exc()

    return diagnostics


def main():
    all_diag = {"docx_headings": [], "sheets": {}}
    try:
        print("Extracting DOCX headings...")
        doc_headings = extract_docx_headings(DOCX_PATH)
        for h in doc_headings:
            print(f" {h['style']}: {h['text']}")
        all_diag["docx_headings"] = doc_headings

        print("\nAnalyzing Excel workbook...")
        xdiag = analyze_excel(XLSX_PATH)
        all_diag["sheets"] = xdiag

        # collect candidates
        candidates = []
        for sheet, entries in xdiag.items():
            for e in entries:
                if e.get("score", 0) >= THRESHOLD:
                    candidates.append({"sheet": sheet, "column": e.get("column"), "score": e.get("score")})

        # save diagnostics JSON
        try:
            OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
            with OUT_JSON.open("w", encoding="utf-8") as f:
                json.dump(all_diag, f, indent=2, ensure_ascii=False)
            print(f"\nDiagnostics saved to: {OUT_JSON}")
        except Exception as e:
            print(f"Failed to write diagnostics JSON: {e}")

        # summary
        print("\nSummary of candidate columns (score >= {0}):".format(THRESHOLD))
        if not candidates:
            print(" None found.")
        else:
            for c in candidates:
                print(f" - Sheet: {c['sheet']}  Column: '{c['column']}'  Score: {c['score']}")

    except Exception as e:
        print(f"Unhandled exception: {e}")
        traceback.print_exc()


if __name__ == '__main__':
    main()
